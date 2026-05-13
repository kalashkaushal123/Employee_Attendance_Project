import mysql.connector
from openpyxl import Workbook
from openpyxl.styles import Font

# Connect MySQL
conn = mysql.connector.connect(
    host = "localhost",
    user = "root",
    password = "1234",
    database = "attendance_db"
)
cursor = conn.cursor()



# Create Excel Workbook
wb = Workbook()

# Remove default sheet
default_sheet = wb.active
wb.remove(default_sheet)



# Separate Sheets
employees_sheet = wb.create_sheet("Employees")
attendance_sheet = wb.create_sheet("Attendance")
leave_sheet = wb.create_sheet("Leave Requests")
shift_sheet = wb.create_sheet("Employee Shifts")




# Employees Sheet

employees_query = """
SELECT
employee_id,
employee_name,
gender,
department,
phone,
email,
join_date

FROM employees
"""

cursor.execute(employees_query)
employees_data = cursor.fetchall()


employees_headers = [
    "Employee ID",
    "Employee Name",
    "Gender",
    "Department",
    "Phone Number",
    "Email",
    "Join Date"
]

employees_sheet.append(employees_headers)

for row in employees_data:
    employees_sheet.append(row)

# bold headers 
for cell in employees_sheet[1]:
    cell.font = Font(bold=True)




# Attendance Sheet

attendance_query = """
SELECT
attendance_id,
employee_id,
attendance_date,
check_in,
check_out,
status

FROM attendance
"""

cursor.execute(attendance_query)
attendance_data = cursor.fetchall()

attendance_headers = [
    "Attendance ID",
    "Employee ID",
    "Attendance Date",
    "Check In Time",
    "Check Out Time",
    "Status"
]

attendance_sheet.append(attendance_headers)

for row in attendance_data:
    attendance_sheet.append(row)

# Bold Headers
for cell in attendance_sheet[1]:
    cell.font = Font(bold=True)






# leave requests Sheet

leave_query = """
SELECT
leave_id,
employee_id,
start_date,
end_date,
reason,
status

FROM leave_requests
"""

cursor.execute(leave_query)
leave_data = cursor.fetchall()


leave_headers = [
    "Leave ID",
    "Employee ID",
    "Start Date",
    "End Date",
    "Reason",
    "Status"
]

leave_sheet.append(leave_headers)

for row in leave_data:
    leave_sheet.append(row)

# Bold Headers
for cell in leave_sheet[1]:
    cell.font = Font(bold=True)





# Employee Shifts Sheet


try:
    shift_query = """
    SELECT
    shift_id,
    employee_id,
    shift_name,
    start_time,
    end_time

    FROM employee_shifts
    """

    cursor.execute(shift_query)
    shift_data = cursor.fetchall()

    shift_headers = [
        "Shift ID",
        "Employee ID"
        "Shift Name",
        "Start Time",
        "End Time"
    ]

    shift_sheet.append(shift_headers)

    for row in shift_data:
        shift_sheet.append(row)

    # Bold Headers
    for cell in shift_sheet[1]:
        cell.font = Font(bold=True)

except:
    print("employee_shifts table not found")


# ws.title = "Attendance Report"


# Auto Column Width

for sheet in wb.worksheets:
    for column in sheet.columns:
        max_length = 0
        column_letter = column[0].column_letter

        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass

        adjusted_width = max_length + 2
        sheet.column_dimensions[column_letter].width = adjusted_width


# Save Excel File
wb.save("../excel_reports/attendance_report.xlsx")

print("Excel report created successfully!")

